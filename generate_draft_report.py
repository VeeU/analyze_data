import os
import re
import shutil
import sys
import tkinter as Tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import time
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score


class battery_cycle_life:

    def issue_draft_report(self,tr_folder_path, root_folder):
        # Dynamically construct the path to the report template by joining root_folder with the relative path
        template_path = os.path.join(root_folder, "report_template", "Battery cycle life.xlsx")

        # Ensure the target folder exists
        target_folder = os.path.join(tr_folder_path, "battery cycle life")
        os.makedirs(target_folder, exist_ok=True)

        # Define the target path for the report template
        target_template_path = os.path.join(target_folder, "Battery cycle life.xlsx")

        try:
            # Copy the report template to the target folder
            shutil.copy2(template_path, target_template_path)  # Use copy2 to preserve metadata
            print(f"Report template copied to {target_template_path}")
        except Exception as e:
            print(f"Error copying report template: {e}")

    import tkinter as Tk
    from tkinter import filedialog, messagebox, simpledialog

    def select_cycle_life_folders(self):
        # Open a dialog to let the user select multiple folders
        root = Tk.Tk()
        root.withdraw()  # Hide the root window

        # List to store the selected folder paths along with milestone values
        folder_data = []

        while True:
            folder_path = filedialog.askdirectory(mustexist=True, title="Select ABCS Cycle Life Folder")
            if folder_path:
                # Prompt for milestone values after each folder selection
                standard_milestone = simpledialog.askinteger(
                    "Input",
                    "Enter the standard milestone (e.g., 250):",
                    initialvalue=250
                )
                reference_milestone = simpledialog.askinteger(
                    "Input",
                    "Enter the reference milestone (e.g., 500):",
                    initialvalue=500
                )
                print(f"Standard Milestone: {standard_milestone}, Reference Milestone: {reference_milestone}")

                # Add folder path and milestone values as a tuple
                folder_data.append((folder_path, standard_milestone, reference_milestone))

                # Ask if the user wants to select another folder
                if not messagebox.askyesno("Select Another?", "Do you want to select another folder?"):
                    break
            else:
                if not folder_data:
                    messagebox.showwarning("Warning", "No folder selected. Exiting...")
                    return None
                break

        return folder_data

    def predict_values_with_evaluation(self,group, y_col, prediction_cycles):
        """Fit a linear regression model, evaluate performance, and predict values for the given cycles."""
        X = group['Cycle #'].values.reshape(-1, 1)
        y = group[y_col].values

        # Split the data into training and testing sets (80% train, 20% test)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # Fit the linear regression model on the training data
        model = LinearRegression().fit(X_train, y_train)

        # Make predictions on the test data
        y_pred = model.predict(X_test)

        # Evaluate the model performance on the test set (only MAE and R²)
        mae = mean_absolute_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)

        # Make predictions for the given prediction cycles (e.g., cycle 250 and 500)
        predictions = model.predict(prediction_cycles)

        # Return predictions for the 250 and 500 cycles, and the evaluation metrics (MAE and R²)
        return predictions[0], predictions[1], mae, r2

    def save_chart_and_table_to_excel(self, df_table, report_path, chart_image_path):

        # Step 1: Load the workbook and insert the chart
        wb = load_workbook(report_path)
        ws = wb['Pack cycle life']

        # Step 2: Find the last row in the worksheet
        start_row = ws.max_row + 2  # Set start_row to the row after the last populated row
        start_col = 1  # Column 'A'

        # Step 3: Insert the DataFrame into the worksheet row by row
        for r_idx, row in enumerate(dataframe_to_rows(df_table, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Step 4: Save the workbook with the inserted table
        #wb.save(report_path)

        time.sleep(2)

        # Calculate the row where the image should be inserted
        image_row = start_row + 6

        # Insert the chart image starting at the calculated position (e.g., A{image_row})
        img = Image(chart_image_path)
        ws.add_image(img, f'A{image_row}')

        time.sleep(2)

        # Save the updated workbook
        try:
            wb.save(report_path)
            print("Workbook saved successfully.")
        except Exception as e:
            print(f"Error saving workbook: {e}")

        time.sleep(1)

    def display_formatted_table(self,df):
        """Display a well-formatted table with actual and predicted values using matplotlib."""
        fig, ax = plt.subplots(figsize=(10, 2))  # Set size based on the data

        # Hide axes
        ax.xaxis.set_visible(False)
        ax.yaxis.set_visible(False)
        ax.set_frame_on(False)

        # Create the table and display it
        table = plt.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.2)  # Scale table to make it more readable

        plt.tight_layout()
        plt.show()

    def draw_charts_with_prediction(self, data, standard_milestone, reference_milestone, chart_image_folder):
        """Draw the charts for Disch. Capacity (AHr), Disch. Energy (WHr), and DCIR (mOhm) with predictions."""
        # Group data by Sample and Model
        groups = data.groupby(['Sample', 'Model'])

        Battery_cycle_life = battery_cycle_life()

        # Create subplots 4for each Y axis data
        fig, axes = plt.subplots(3, 1, figsize=(10, 15), sharex=True)

        # Define prediction points using user-defined milestones
        prediction_cycles = np.array([standard_milestone, reference_milestone]).reshape(-1, 1)

        # Iterate through each group and plot the values
        for (sample, model), group in groups:
            # Plot and predict Disch. Capacity (AHr)
            Battery_cycle_life.plot_with_prediction(axes[0], group, 'Disch. Capacity (AHr)', sample, model, prediction_cycles)
            axes[0].set_ylabel('Disch. Capacity (AHr)')
            axes[0].set_title(
                f'Disch. Capacity (AHr) vs Cycle # (Milestones: {standard_milestone}, {reference_milestone})')

            # Plot and predict Disch. Energy (WHr)
            Battery_cycle_life.plot_with_prediction(axes[1], group, 'Disch. Energy (WHr)', sample, model, prediction_cycles)
            axes[1].set_ylabel('Disch. Energy (WHr)')
            axes[1].set_title(
                f'Disch. Energy (WHr) vs Cycle # (Milestones: {standard_milestone}, {reference_milestone})')

            # Plot and predict DCIR (mOhm)
            Battery_cycle_life.plot_with_prediction(axes[2], group, 'DCIR (mOhm)', sample, model, prediction_cycles)
            axes[2].set_ylabel('DCIR (mOhm)')
            axes[2].set_title(
                f'DCIR (mOhm) vs Cycle # (Milestones: {standard_milestone}, {reference_milestone})')

        # Set x-axis label for the last plot
        axes[2].set_xlabel('Cycle #')

        # Add legends
        for ax in axes:
            ax.legend()

        # Display the plot
        plt.tight_layout()
        #plt.close(fig)

        time.sleep(1)

        chart_image_path = os.path.join(chart_image_folder, 'chart.png')  # Full path for the saved image
        plt.savefig(chart_image_path)

        # Step 3: Return the file path to the saved chart
        return chart_image_path

    def plot_with_prediction(self,ax, group, y_col, sample, model, prediction_cycles):
        """Fit a linear regression model, plot the data, and make predictions."""
        # Prepare data for linear regression
        X = group['Cycle #'].values.reshape(-1, 1)
        y = group[y_col].values

        # Plot the actual data
        ax.plot(X, y, label=f'{sample} - {model} Actual')

        # Fit the linear regression model
        linear_model = LinearRegression().fit(X, y)

        # Predict the values for the user-specified milestones
        predictions = linear_model.predict(prediction_cycles)

        # Plot the predicted values at the specified milestones
        ax.scatter(prediction_cycles, predictions, color='red', marker='x', label=f'{sample} - {model} Prediction')

        # Annotate the predictions with milestone labels
        for i, cycle in enumerate(prediction_cycles):
            ax.text(cycle, predictions[i], f'{predictions[i]:.2f} @ Cycle {cycle[0]}', fontsize=10, color='red')

    def run_analysis(self, parent_folder_path, standard_milestone, reference_milestone, tr_folder_path):
        """Combines folder browsing, data extraction, outlier removal, prediction, and charting in one method."""
        Battery_cycle_life = battery_cycle_life()

        child_folders = []

        # Loop over each selected parent folder path
        print('hi')

        # Iterate over all items in the parent folder
        for item in os.listdir(parent_folder_path):
            item_path = os.path.join(parent_folder_path, item)
            # Check if the item is a directory
            if os.path.isdir(item_path):
                child_folders.append(item_path)

        # Create a DataFrame to collect all data
        all_data = pd.DataFrame()

        # Loop through each child folder and look for data.csv
        for child_folder in child_folders:
            data_file_path = os.path.join(child_folder, "data.csv")
            if os.path.exists(data_file_path):
                try:
                    # Read the CSV file and extract specific columns
                    df = pd.read_csv(data_file_path,
                                     usecols=['Cycle #', 'Disch. Capacity (AHr)', 'Disch. Energy (WHr)', 'DCIR (mOhm)'])
                    df['Sample'] = os.path.basename(child_folder)
                    df['Model'] = os.path.basename(os.path.dirname(child_folder))
                    # Remove outliers
                    for col in ['Disch. Capacity (AHr)', 'Disch. Energy (WHr)', 'DCIR (mOhm)']:
                        Q1 = df[col].quantile(0.25)
                        Q3 = df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        df = df[(df[col] >= (Q1 - 1.5 * IQR)) & (df[col] <= (Q3 + 1.5 * IQR))]

                    # Append cleaned data
                    all_data = pd.concat([all_data, df], ignore_index=True)
                except Exception as e:
                    print(f"Error processing file {data_file_path}: {e}")
                    continue

        # Save or append
        abusive_file_path = os.path.join(tr_folder_path, "battery cycle life", "cycle.csv")
        if not all_data.empty:
            if os.path.exists(abusive_file_path):
                all_data.to_csv(abusive_file_path, mode='a', header=False, index=False)
                print(f"Appended data to {abusive_file_path}")
            else:
                all_data.to_csv(abusive_file_path, mode='w', header=True, index=False)
                print(f"Created and saved data to {abusive_file_path}")

            # Prepare data for predictions
            prediction_cycles = np.array([standard_milestone, reference_milestone]).reshape(-1, 1)
            groups = all_data.groupby(['Sample', 'Model'])

            table_data = []
            for (sample, model), group in groups:
                initial_data = group[(group['Cycle #'] >= 6) & (group['Cycle #'] <= 10)].mean()
                last_cycle_data = group.iloc[-1]

                # Predict values and calculate retention
                capacity_pred_std, capacity_pred_ref, mae_capacity, r2_capacity = Battery_cycle_life.predict_values_with_evaluation(group,
                                                                                               'Disch. Capacity (AHr)',
                                                                                               prediction_cycles)
                energy_pred_std, energy_pred_ref, mae_energy, r2_energy = Battery_cycle_life.predict_values_with_evaluation(group, 'Disch. Energy (WHr)',
                                                                                           prediction_cycles)
                dcir_pred_std, dcir_pred_ref, mae_dcir, r2_dcir = Battery_cycle_life.predict_values_with_evaluation(group, 'DCIR (mOhm)',
                                                                                       prediction_cycles)

                table_data.append({
                    'Sample': sample, 'Model': model,
                    'Initial Capacity (6-10 avg)': round(initial_data['Disch. Capacity (AHr)'], 2),
                    'Initial Energy (6-10 avg)': round(initial_data['Disch. Energy (WHr)'], 2),
                    'Initial DCIR (6-10 avg)': round(initial_data['DCIR (mOhm)'], 2),
                    'Last cycle': last_cycle_data['Cycle #'],
                    'Capacity Last': round(last_cycle_data['Disch. Capacity (AHr)'], 2),
                    'Energy Last': round(last_cycle_data['Disch. Energy (WHr)'], 2),
                    'DCIR Last': round(last_cycle_data['DCIR (mOhm)'], 2),
                    'Capacity Last Retention (%)': round(last_cycle_data['Disch. Capacity (AHr)']*100/round(initial_data['Disch. Capacity (AHr)'], 2), 2),
                    'Energy Last Retention (%)': round(last_cycle_data['Disch. Energy (WHr)'] * 100 / round(initial_data['Disch. Energy (WHr)'],2), 2),
                    'DCIR Last Increase (%)': round(((last_cycle_data['DCIR (mOhm)']/initial_data['DCIR (mOhm)'])-1)*100,2),
                    f'Capacity Pred {standard_milestone}': round(capacity_pred_std, 2),
                    f'Energy Pred {standard_milestone}': round(energy_pred_std, 2),
                    f'DCIR Pred {standard_milestone}': round(dcir_pred_std, 2),

                    'Capacity Retention 250(%)': round(capacity_pred_std * 100 / initial_data['Disch. Capacity (AHr)'], 2),
                    'Energy Retention 250(%)': round(energy_pred_std * 100 / initial_data['Disch. Energy (WHr)'],2),
                    'DCIR Increase 250(%)': round(((dcir_pred_std / initial_data['DCIR (mOhm)']) - 1) * 100, 2),

                    f'Capacity Pred {reference_milestone}': round(capacity_pred_ref, 2),
                    f'Energy Pred {reference_milestone}': round(energy_pred_ref, 2),
                    f'DCIR Pred {reference_milestone}': round(dcir_pred_ref, 2),

                    'Capacity Retention 500(%)': round(capacity_pred_ref * 100 / initial_data['Disch. Capacity (AHr)'],2),
                    'Energy Retention 500(%)': round(energy_pred_ref * 100 / initial_data['Disch. Energy (WHr)'], 2),
                    'DCIR Increase 500(%)': round(((dcir_pred_ref / initial_data['DCIR (mOhm)']) - 1) * 100, 2),

                    'Capacity MAE': round(mae_capacity,2),
                    'Capacity R²': round(r2_capacity,2),
                    'Energy MAE': round(mae_energy,2),
                    'Energy R²': round(r2_energy,2),
                    'DCIR MAE': round(mae_dcir,2),
                    'DCIR R²': round(r2_dcir,2),
                    # Similar entries for Energy and DCIR
                })

            df_table = pd.DataFrame(table_data)

            # Draw charts with predictions
            chart_image_folder = os.path.join(tr_folder_path, "battery cycle life")
            Battery_cycle_life.draw_charts_with_prediction(all_data, standard_milestone, reference_milestone, chart_image_folder)

            report_path = os.path.join(tr_folder_path, "battery cycle life", "Battery cycle life.xlsx")
            Battery_cycle_life.save_chart_and_table_to_excel(df_table,report_path, os.path.join(tr_folder_path, "battery cycle life", "chart.png"))
            Battery_cycle_life.display_formatted_table(df_table)



        else:
            print("No valid data found to append.")

class pack_discharge:
    @staticmethod
    def select_discharge_folders():
        # Open a dialog to let the user select multiple folders
        root = Tk.Tk()
        root.withdraw()  # Hide the root window
        folder_paths = filedialog.askdirectory(mustexist=True, title="Select ABCS Pack Discharge Folder")
        if not folder_paths:
            messagebox.showwarning("Warning", "No folder selected. Exiting...")
            return None
        return folder_paths

    @staticmethod
    def copy_folders_to_tr_path(tr_folder_path, discharge_folders):
        # Target folder where files will be copied
        target_folder = os.path.join(tr_folder_path, "battery pack discharge")
        os.makedirs(target_folder, exist_ok=True)  # Create target folder if it doesn't exist

        # Loop through all selected discharge folders and copy their entire contents, including the folder structure itself
        for folder in discharge_folders:
            if os.path.exists(folder):
                # Create the target directory within the target folder (preserve the discharge folder name)
                folder_name = os.path.basename(folder)
                target_discharge_folder = os.path.join(target_folder, folder_name)
                os.makedirs(target_discharge_folder, exist_ok=True)

                # Recursively copy all files and folders from each discharge folder to target discharge folder
                for root, dirs, files in os.walk(folder):
                    # Construct the target path by maintaining the directory structure
                    relative_path = os.path.relpath(root, folder)
                    target_dir = os.path.join(target_discharge_folder, relative_path)
                    os.makedirs(target_dir, exist_ok=True)

                    # Copy each file from source to target directory
                    for file in files:
                        source_file = os.path.join(root, file)
                        target_file = os.path.join(target_dir, file)

                        try:
                            shutil.copy2(source_file, target_file)  # Use copy2 to preserve metadata
                            print(f"Copied {source_file} to {target_file}")
                        except Exception as e:
                            print(f"Error copying {source_file} to {target_file}: {e}")
            else:
                print(f"Folder {folder} does not exist.")

    @staticmethod
    def issue_draft_report(tr_folder_path, root_folder):
        # Dynamically construct the path to the report template by joining root_folder with the relative path
        template_path = os.path.join(root_folder, "report_template", "Battery pack discharge.xlsx")

        # Ensure the target folder exists
        target_folder = os.path.join(tr_folder_path, "battery pack discharge")
        os.makedirs(target_folder, exist_ok=True)

        # Define the target path for the report template
        target_template_path = os.path.join(target_folder, "Battery pack discharge.xlsx")

        try:
            # Copy the report template to the target folder
            shutil.copy2(template_path, target_template_path)  # Use copy2 to preserve metadata
            print(f"Report template copied to {target_template_path}")
        except Exception as e:
            print(f"Error copying report template: {e}")

    @staticmethod
    def scan_and_copy_csv_to_excel(folders, excel_template_path):
        # Load the Excel workbook and select the active sheet
        headers = [
            'Sample #', 'Test condition', 'Charge Time (min)', 'Discharge Time (min)',
            'Ch. Capacity (AHr)', 'Ch. Energy (WHr)', 'Disch. Capacity (AHr)', 'Disch. Energy (WHr)',
            'Ch. Stop Voltage', 'Disch. Stop Voltage', 'Disch. Stop Temp', 'Disch. Stop Type',
            'DCIR (mOhm)', 'Disch. End Time', 'Rebound voltage'
        ]

        workbook = load_workbook(excel_template_path)
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(excel_template_path)

        # Create an empty DataFrame with the specified headers
        df = pd.DataFrame()

        # Loop through each selected folder
        for folder in folders:

            # Extract the folder name and append it to the Excel sheet
            folder_name = os.path.basename(folder)

            # Regular expression to capture FW, HW, and Cdegree parts
            fw_match = re.search(r'FW[\d.]+', folder_name)
            hw_match = re.search(r'HW[\d.]+', folder_name)
            cdegree_match = re.search(r'(-?\s*\d+)Cdegree', folder_name)

            # Extract matched values or set as empty if not found
            fw = fw_match.group(0) if fw_match else ''
            hw = hw_match.group(0) if hw_match else ''
            cdegree = cdegree_match.group(0) if cdegree_match else ''

            # Merge the extracted parts with hyphens
            test_condition = '-'.join(filter(None, [fw, hw, cdegree]))

            # Scan for 'data.csv' in level-1 subfolders
            for subfolder in os.listdir(folder):
                subfolder_path = os.path.join(folder, subfolder)

                if os.path.isdir(subfolder_path):
                    data_csv_path = os.path.join(subfolder_path, 'data.csv')
                    if os.path.exists(data_csv_path):
                        print(f"Found 'data.csv' in {data_csv_path}")

                        # Read the CSV file into a DataFrame, including the headers
                        df_new = pd.read_csv(data_csv_path)

                        # Replace the first column of the DataFrame (Cycle #) with the subfolder name
                        df_new.iloc[:, 0] = subfolder

                        # Step 1: Extract the second column ('Test condition')
                        original_second_column = df_new['Discharge Current (A)']

                        # Step 2: Create a new second column by joining test_condition with the original data
                        new_second_column = original_second_column.apply(lambda x: f"{test_condition}-{x}A")

                        # Step 3: Drop the original second column
                        df_new.drop('Discharge Current (A)', axis=1, inplace=True)

                        # Step 4: Insert the new second column back at the same position
                        df_new.insert(1, 'Discharge Current (A)', new_second_column)

                        # Append the current DataFrame to the accumulated DataFrame
                        df = pd.concat([df_new, df], ignore_index=True)

            # Sort the DataFrame by the second column ('Discharge Current (A)')
            df_sorted = df.sort_values(by=[df.columns[1], df.columns[0]], ascending=[True, True])

        # Append the DataFrame contents row by row
        for row in df_sorted.itertuples(index=False, name=None):
            sheet.append(row)

        # Save the workbook with the new data added
        workbook.save(excel_template_path)
        print(
            f"Appended data from {data_csv_path} to Excel template with headers and folder name '{folder_name}'.")

        # Delay for 1 second to ensure the save action is completed
        time.sleep(1)

        #Chart Ahr
        # Pivot the data to show 'Disch. Capacity (AHr)' values with 'Cycle #' as columns and 'Discharge Current (A)' as rows
        analyze_Ah_dataframe = df_sorted.pivot(index="Discharge Current (A)", columns="Cycle #",
                                            values="Disch. Capacity (AHr)")

        # Create a bar chart from the transposed pivoted data
        plt.figure(figsize=(12, 8))
        analyze_Ah_dataframe.plot(kind='bar')

        # Set titles and labels
        plt.title('Disch. Capacity (AHr) by Test condition and Sample #', fontsize=14)
        plt.xlabel('Test condition', fontsize=12)
        plt.ylabel('Disch. Capacity (AHr)', fontsize=12)

        # Customize the legend and grid
        plt.legend(title='Sample #', bbox_to_anchor=(1.05, 1), loc='upper left')  # Place legend outside the plot
        plt.grid(True)

        # Adjust layout to avoid overlap
        plt.tight_layout()

        # Save the plot as an image
        image_path = os.path.join(os.path.dirname(excel_template_path), "chart.png")
        plt.savefig(image_path)

        # Insert the image into the Excel template
        img = Image(image_path)
        sheet.add_image(img, 'Q18')  # Adjust the cell location as needed

        # Save the Excel file with the embedded image
        workbook.save(excel_template_path)

        # Final 1-second delay after the last save
        time.sleep(1)

        # Show the plot
        plt.show()

        # Save the updated Excel file again after processing all folders
        workbook.save(excel_template_path)

    @staticmethod
    def select_multiple_folders():
        root = Tk.Tk()
        root.withdraw()  # Hide the root window

        selected_folders = []
        while True:
            folder_path = filedialog.askdirectory(mustexist=True, title="Select a Folder (Cancel to finish)")
            if not folder_path:
                break  # Exit if the user cancels or doesn't select a folder
            selected_folders.append(folder_path)

            # Ask if the user wants to select another folder
            if not messagebox.askyesno("Confirmation", "Do you want to select another folder?"):
                break  # Stop asking for more folders if the user says 'No'

        return selected_folders

class software_verification:
    @staticmethod
    def select_discharge_folders():
        # Open a dialog to let the user select multiple folders
        root = Tk.Tk()
        root.withdraw()  # Hide the root window
        folder_paths = filedialog.askdirectory(mustexist=True, title="Select ABCS Pack Discharge Folder")
        if not folder_paths:
            messagebox.showwarning("Warning", "No folder selected. Exiting...")
            return None
        return folder_paths

    @staticmethod
    def copy_folders_to_tr_path(tr_folder_path, discharge_folders):
        # Target folder where files will be copied
        target_folder = os.path.join(tr_folder_path, "battery pack discharge")
        os.makedirs(target_folder, exist_ok=True)  # Create target folder if it doesn't exist

        # Loop through all selected discharge folders and copy their entire contents, including the folder structure itself
        for folder in discharge_folders:
            if os.path.exists(folder):
                # Create the target directory within the target folder (preserve the discharge folder name)
                folder_name = os.path.basename(folder)
                target_discharge_folder = os.path.join(target_folder, folder_name)
                os.makedirs(target_discharge_folder, exist_ok=True)

                # Recursively copy all files and folders from each discharge folder to target discharge folder
                for root, dirs, files in os.walk(folder):
                    # Construct the target path by maintaining the directory structure
                    relative_path = os.path.relpath(root, folder)
                    target_dir = os.path.join(target_discharge_folder, relative_path)
                    os.makedirs(target_dir, exist_ok=True)

                    # Copy each file from source to target directory
                    for file in files:
                        source_file = os.path.join(root, file)
                        target_file = os.path.join(target_dir, file)

                        try:
                            shutil.copy2(source_file, target_file)  # Use copy2 to preserve metadata
                            print(f"Copied {source_file} to {target_file}")
                        except Exception as e:
                            print(f"Error copying {source_file} to {target_file}: {e}")
            else:
                print(f"Folder {folder} does not exist.")

    @staticmethod
    def issue_draft_report(tr_folder_path, root_folder):
        # Dynamically construct the path to the report template by joining root_folder with the relative path
        template_path = os.path.join(root_folder, "report_template", "Battery pack discharge.xlsx")

        # Ensure the target folder exists
        target_folder = os.path.join(tr_folder_path, "battery pack discharge")
        os.makedirs(target_folder, exist_ok=True)

        # Define the target path for the report template
        target_template_path = os.path.join(target_folder, "Battery pack discharge.xlsx")

        try:
            # Copy the report template to the target folder
            shutil.copy2(template_path, target_template_path)  # Use copy2 to preserve metadata
            print(f"Report template copied to {target_template_path}")
        except Exception as e:
            print(f"Error copying report template: {e}")

    @staticmethod
    def scan_and_copy_csv_to_excel(folders, excel_template_path):
        # Load the Excel workbook and select the active sheet
        headers = [
            'Sample #', 'Test condition', 'Charge Time (min)', 'Discharge Time (min)',
            'Ch. Capacity (AHr)', 'Ch. Energy (WHr)', 'Disch. Capacity (AHr)', 'Disch. Energy (WHr)',
            'Ch. Stop Voltage', 'Disch. Stop Voltage', 'Disch. Stop Temp', 'Disch. Stop Type',
            'DCIR (mOhm)', 'Disch. End Time', 'Rebound voltage'
        ]

        workbook = load_workbook(excel_template_path)
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(excel_template_path)

        # Create an empty DataFrame with the specified headers
        df = pd.DataFrame()

        # Loop through each selected folder
        for folder in folders:

            # Extract the folder name and append it to the Excel sheet
            folder_name = os.path.basename(folder)

            # Regular expression to capture FW, HW, and Cdegree parts
            fw_match = re.search(r'FW[\d.]+', folder_name)
            hw_match = re.search(r'HW[\d.]+', folder_name)
            cdegree_match = re.search(r'(-?\s*\d+)Cdegree', folder_name)

            # Extract matched values or set as empty if not found
            fw = fw_match.group(0) if fw_match else ''
            hw = hw_match.group(0) if hw_match else ''
            cdegree = cdegree_match.group(0) if cdegree_match else ''

            # Merge the extracted parts with hyphens
            test_condition = '-'.join(filter(None, [fw, hw, cdegree]))

            # Scan for 'data.csv' in level-1 subfolders
            for subfolder in os.listdir(folder):
                subfolder_path = os.path.join(folder, subfolder)

                if os.path.isdir(subfolder_path):
                    data_csv_path = os.path.join(subfolder_path, 'data.csv')
                    if os.path.exists(data_csv_path):
                        print(f"Found 'data.csv' in {data_csv_path}")

                        # Read the CSV file into a DataFrame, including the headers
                        df_new = pd.read_csv(data_csv_path)

                        # Replace the first column of the DataFrame (Cycle #) with the subfolder name
                        df_new.iloc[:, 0] = subfolder

                        # Step 1: Extract the second column ('Test condition')
                        original_second_column = df_new['Discharge Current (A)']

                        # Step 2: Create a new second column by joining test_condition with the original data
                        new_second_column = original_second_column.apply(lambda x: f"{test_condition}-{x}A")

                        # Step 3: Drop the original second column
                        df_new.drop('Discharge Current (A)', axis=1, inplace=True)

                        # Step 4: Insert the new second column back at the same position
                        df_new.insert(1, 'Discharge Current (A)', new_second_column)

                        # Append the current DataFrame to the accumulated DataFrame
                        df = pd.concat([df_new, df], ignore_index=True)

            # Sort the DataFrame by the second column ('Discharge Current (A)')
            df_sorted = df.sort_values(by=[df.columns[1], df.columns[0]], ascending=[True, True])

        # Append the DataFrame contents row by row
        for row in df_sorted.itertuples(index=False, name=None):
            sheet.append(row)

        # Save the workbook with the new data added
        workbook.save(excel_template_path)
        print(
            f"Appended data from {data_csv_path} to Excel template with headers and folder name '{folder_name}'.")

        # Delay for 1 second to ensure the save action is completed
        time.sleep(1)

        #Chart Ahr
        # Pivot the data to show 'Disch. Capacity (AHr)' values with 'Cycle #' as columns and 'Discharge Current (A)' as rows
        analyze_Ah_dataframe = df_sorted.pivot(index="Discharge Current (A)", columns="Cycle #",
                                            values="Disch. Capacity (AHr)")

        # Create a bar chart from the transposed pivoted data
        plt.figure(figsize=(12, 8))
        analyze_Ah_dataframe.plot(kind='bar')

        # Set titles and labels
        plt.title('Disch. Capacity (AHr) by Test condition and Sample #', fontsize=14)
        plt.xlabel('Test condition', fontsize=12)
        plt.ylabel('Disch. Capacity (AHr)', fontsize=12)

        # Customize the legend and grid
        plt.legend(title='Sample #', bbox_to_anchor=(1.05, 1), loc='upper left')  # Place legend outside the plot
        plt.grid(True)

        # Adjust layout to avoid overlap
        plt.tight_layout()

        # Save the plot as an image
        image_path = os.path.join(os.path.dirname(excel_template_path), "chart.png")
        plt.savefig(image_path)

        # Insert the image into the Excel template
        img = Image(image_path)
        sheet.add_image(img, 'Q18')  # Adjust the cell location as needed

        # Save the Excel file with the embedded image
        workbook.save(excel_template_path)

        # Final 1-second delay after the last save
        time.sleep(1)

        # Show the plot
        plt.show()

        # Save the updated Excel file again after processing all folders
        workbook.save(excel_template_path)

    @staticmethod
    def select_multiple_folders():
        root = Tk.Tk()
        root.withdraw()  # Hide the root window

        selected_folders = []
        while True:
            folder_path = filedialog.askdirectory(mustexist=True, title="Select a Folder (Cancel to finish)")
            if not folder_path:
                break  # Exit if the user cancels or doesn't select a folder
            selected_folders.append(folder_path)

            # Ask if the user wants to select another folder
            if not messagebox.askyesno("Confirmation", "Do you want to select another folder?"):
                break  # Stop asking for more folders if the user says 'No'

        return selected_folders

def check_file_value(file_path, expected_value):
    try:
        with open(file_path, 'r') as file:
            value = file.read().strip()
            return value == expected_value
    except:
        return False

# Main execution logic
if __name__ == "__main__":
    # Path to the file you want to check
    file_path = "V:\\AESVN\\Public\\Tom\\text.txt"

    # Check the value of the file before proceeding
    if not check_file_value(file_path, "1"):
        sys.exit(0)  # Exit the program if the file value is not "1"

    # Check if the three arguments (rootFolder, trFolderPath, and selectedTestItem) are passed
    if len(sys.argv) < 4:
        print("Insufficient arguments provided.")
        sys.exit(1)

    # Extract the arguments from the command line
    root_folder = sys.argv[1]
    tr_folder_path = sys.argv[2]
    selected_test_item = sys.argv[3]

    #root_folder = "V:/NPI/Share/1.LAB OP/9. CNB Test Report/raw"
    #tr_folder_path = "V:/NPI/Share/1.LAB OP/9. CNB Test Report/raw/test_report/CNB-20240803-OP40806T-130518001DG9"
    #selected_test_item = "battery cycle life"

    if selected_test_item == "battery pack discharge":
        # Display the passed arguments as a pop-up notification using tkinter
        root = Tk.Tk()
        root.withdraw()  # Hide the root window
        messagebox.showinfo("Passed Arguments", f"Root Folder: {root_folder}\nTR Folder Path: {tr_folder_path}\nSelected Test Item: {selected_test_item}")

        # Step 1: Copy report template to trFolderPath/pack discharge
        pack_discharge.issue_draft_report(tr_folder_path, root_folder)

        # Step 2: Let the user select multiple pack discharge folders from ABCS
        folders = pack_discharge.select_multiple_folders()
        if folders:
            # Join the folder paths into a string to display in a message box
            folders_str = "\n".join(folders)
            messagebox.showinfo("Selected Folders", f"The following folders were selected:\n\n{folders_str}")
        else:
            messagebox.showinfo("No Folders Selected", "No folders were selected.")
            sys.exit(1)

        # Step 3: Scan for data.csv and append its contents to the Excel template
        report_path = os.path.join(tr_folder_path, "battery pack discharge", "Battery pack discharge.xlsx")
        pack_discharge.scan_and_copy_csv_to_excel(folders, report_path)

        messagebox.showinfo("Process Complete", "Data has been copied to the report template and saved.")

    if selected_test_item == "battery cycle life":
        # Step 1: Copy report template to trFolderPath/battery cycle life
        Battery_cycle_life = battery_cycle_life()
        Battery_cycle_life.issue_draft_report(tr_folder_path, root_folder)

        # Step 2: Let the user select multiple battery cycle life folders from ABCS
        folder_datas = Battery_cycle_life.select_cycle_life_folders()

        # Step 3: Scan for every condition in cycle life and append its contents to the Excel template
        for folder_data in folder_datas:
            print('2')
            print(folder_data[0])
            print(folder_data[1])
            print(folder_data[2])
            Battery_cycle_life.run_analysis(folder_data[0], folder_data[1], folder_data[2], tr_folder_path)
