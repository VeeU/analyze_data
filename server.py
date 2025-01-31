import discord
import pandas as pd
import os
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Define the bot client with intents
intents = discord.Intents.default()
intents.messages = True
client = discord.Client(intents=intents)

# Excel file path
EXCEL_FILE_PATH = r'C:\Users\aesavle\OneDrive - Techtronic Industries Co. Ltd\Work\5. Task_sync\2022\8_app_db\data\raw\test_report\CNB-20240802-OP40406T-130805002DG9\discord_messages.xlsx'

# Image save folder
IMAGE_SAVE_FOLDER = r'C:\Users\aesavle\OneDrive - Techtronic Industries Co. Ltd\Work\5. Task_sync\2022\8_app_db\data\raw\test_report\CNB-20240802-OP40406T-130805002DG9\images'


# Ensure the Excel file and image folder exist, if not, create them
def ensure_file_and_folder():
    # Create the image folder if it doesn't exist
    if not os.path.exists(IMAGE_SAVE_FOLDER):
        os.makedirs(IMAGE_SAVE_FOLDER)
        print(f"Created image folder: {IMAGE_SAVE_FOLDER}")

    # Create Excel file if it doesn't exist
    if not os.path.exists(EXCEL_FILE_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['User', 'Message', 'Image'])  # Add header
        workbook.save(EXCEL_FILE_PATH)
        print(f"Created Excel file: {EXCEL_FILE_PATH}")


# Function to save messages and images to Excel
def save_message_to_excel(username, message_text, image_filename=None):
    try:
        # Load the workbook and sheet
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active

        # Append text message to the next row
        row = [username, message_text]
        sheet.append(row)
        row_num = sheet.max_row

        # If there's an image, add it to the same row
        if image_filename:
            img = ExcelImage(image_filename)
            img.width, img.height = 100, 100  # Resize image
            img_cell = f"C{row_num}"  # Column C for the image
            sheet.add_image(img, img_cell)

        # Save the Excel file
        workbook.save(EXCEL_FILE_PATH)

    except Exception as e:
        print(f"Error saving message to Excel: {e}")


# Function to save image locally
def save_image(attachment):
    image_url = attachment.url
    image_filename = os.path.join(IMAGE_SAVE_FOLDER, attachment.filename)

    try:
        response = requests.get(image_url)
        if response.status_code == 200:
            with open(image_filename, 'wb') as f:
                f.write(response.content)
            return image_filename
        else:
            print(f"Failed to download image: {response.status_code}")
            return None
    except Exception as e:
        print(f"Error saving image: {e}")
        return None


# Event when the bot has connected to the server
@client.event
async def on_ready():
    print(f'We have logged in as {client.user}')
    ensure_file_and_folder()  # Ensure file and folder exist on bot startup


# Event when a message is sent
@client.event
async def on_message(message):
    # Ignore messages from the bot itself
    if message.author == client.user:
        return

    username = message.author.name
    message_text = message.content

    # Check if there's an image attachment
    if message.attachments:
        for attachment in message.attachments:
            if attachment.content_type.startswith('image/'):  # Check if the attachment is an image
                image_filename = save_image(attachment)  # Save the image
                save_message_to_excel(username, message_text, image_filename)  # Log the image in the Excel file
                await message.channel.send(
                    f"Thanks, {username}! Your message and image have been saved to the Excel file.")
            else:
                save_message_to_excel(username, message_text)  # Save message without an image
                await message.channel.send(
                    f"Thanks, {username}! Your message has been saved, but the attachment was not an image.")
    else:
        save_message_to_excel(username, message_text)  # Save message without any attachment
        await message.channel.send(f"Thanks, {username}! Your message has been saved to the Excel file.")

# Run the bot using the token from Discord Developer Portal
client.run('MTI5MjY5NzYwMDM0ODI2MjQzMA.GbmEDn.IGSaYUU1u0SRUzbK45s1hL8b2FwwMs8czba1z8')
